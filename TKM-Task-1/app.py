from flask import Flask, render_template, render_template_string, session, redirect, url_for, request, send_file
import psycopg2
import psycopg2.extras
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import json

import pandas as pd
import io
import base64, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "tkm_ntf_secret_2025"

PER_PAGE = 10

# Toyota logo embedded directly — no external file needed
TOYOTA_LOGO = "data:image/png;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAFsAcMDASIAAhEBAxEB/8QAHQABAAIDAQEBAQAAAAAAAAAAAAYHBAUIAwIBCf/EAE4QAAEEAQIDBAcEBQgIBQQDAAEAAgMEBQYRBxIhEzFBUQgUImFxgZEyQqGxFSNSYsEWJDM0Q4KS0QkXU3KTosLhVWODsvAlVHPxNURk/8QAFQEBAQAAAAAAAAAAAAAAAAAAAAH/xAAXEQEBAQEAAAAAAAAAAAAAAAAAEQEh/9oADAMBAAIRAxEAPwDstERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBF5WrNevEZrU8UEY73yPDWj5lQLVXGXh9p57oJ87HbsjoIabXTFx8t2Aj8UFhIqJscc9QZeTstGcPMpbJ6CS43s2H39Cvnf0g9Sg7T4zTteQdWtY17mj3O23QXtJJHE3mke1jfNx2C1WT1Pp7GM57+ZowNHi6YfwVOf6ktYZZp/lJxKzFhj/ALcMdh4Z9N9lsMV6NuhK0nbXRavSnq50rydz9UEoyPGnhpTJb/KipYeO9kO7iFpLnpD8P4dxCMvad3Dsam4J/wASkON4RcPqDQItOU3beLowVv6ekdMUgBWwtGPbyhb/AJIKrm9IWpJ1xujs1aG24Lm8nX6FY/8Ar51JI4CrwvyMoI7za5f+hXW2njoRtHTrtHujAXy/1dv2YYx8GhWCkzxh4kWJi6rwzljiPc19ncj58q+pOKPFiWMiDh9DE7wc+wSPpyq45Joh4MH0WNLYhHeWJBUcPEvi7E7msaHqyt27ObflBtMLQiiaGQQRRKOmUe6NvvWF6wRRxNDIY2xRjuaxoAH4KM1uO+mXAt/SUrgfCVw/yCwHcaMiC7TupHiIz/MBBpLurcJVcW/pSkwf7ORyfwXON7XcLJhzPMeRoMzz/JWWx+PbH6oHiWk/NaXJ8IuHGROzsFBA4/3aTon+YCqnUHCrBWy51G7Yxkp7zE/na33A9PogmuF1ppjPRNdj8xVkefuB+3uPzWxXH2o+GV3Eu/S+l3vhsD+2rHciD7wfNQavxfxR0dq1vJj9V4m0T0DWzt3/BTiqL0fxfytF7cLxBoWLkJ6NsVx+8HzB8VeGnOLujNRFrI8oyleY9621IXt/FBN0REBERAREQEREBERAREQEREBERAREQEREBfj3NY0ve4NaBuST0C0OtNXYPSWMfdzFxkWw9iIHd8h8gFVHJrzi7Me1dLpvSrj0jaT21hvmT023QSLWfGGjVvuwWjqT9Q5knl5YAXRRnzc4dPxWno8MNUa1sR5PiZmpJIt+ZmMrO5ImeQO3f81ZOi9F4DSVBtXE0mMdt7crhu958yVI+5Bq9O6ew2n6TKmIx1epE0bewwAn4nvW0PRY925Wp132LM0cMTBu573bAKnNdcc6NaxJjNI0n5e6Dy9r3RtP8AFBcdy7WpwOntTxwRN73yODWj5lVhq7jhpTEzOqY90uYudwjrDdu/xCp3LR6o1XY9d1dmpWsP2alckNA8l904OHvW+hibrQR0OPIAG7/BVjqPjNorFOdHHkpcjMO6OpGXbn4nZQjI8bM/d3GC0g9rT9mW3IB+KC/LnG3RbdxDYyds/wDlxbA/gtdLxpxL/wCraXzNjyLn7b/8qqiu7W9rYUsLYYCdhyVS38192q2pMdD6zqnOxYGA/ZZNJzSye5rGbkn47ILKk4u23/1bQVpw85JXLxfxXzJ+zoJo/wB6VceBfDfS3EjT8eRwl+SKaM8l2jOwNmgf4HqD5EFVGkNKaa0tSFTS+Cq46Edy2GM8/iT3k/FBaA1Zi3qF2LKHuM8DQPzXwNfaecNxlI/8e3+hXSKIObYMTdoNbsRvv0Xz+j8b/4fV/4bf8lZCID22OqMIIPqsfz2/wAl7p6PxcLg+PHVAQdwREwEfkqaRBWWouEuiNSMLrGJjhmd3yxNMZPv22P1Wmy/owxQPExer7gHggmYHD/XZbjZWa0mW7MkDh5H/wBlXGpNG6a1C3bJ0IJpO5+wDXD5j+KC2cDqKfG2Q+vZexn3mMJ5fzUlhmZPG2SM7td1BX8+IxVqoW2oYZWbbb8oIXrT0lioAPVqr4R4MsHl8EFKM09iW/R+OqMHgyJrD+QWhye34fr6zXkePR7mbjb3B239/c+5bsaQx0bbQqNiHSaEBgB9wHRRHU2jYpInuosAlHVpPcfJBjjWlkdCCb4aTxYGy57A+bB1P4hSfS+VxWZom7iy37MpAkY4cyQSNwQe4giql1XUq0bZLENiF0kfOPYe17fMFRlmkqtaS1NIY42Alcep2AP/AJKiulGq3Yba7Zyj90MhH4AoLgOoKoHarXHd5eWrB9ZBQPWGp2Y7S0mgpTk/asDiQPkVhUbmtNMUeX1nOVOb9kScx+AKBT6gzF+VxpMrVAP2Yy0ED3uJ3K9qde7fnEcNbszOO5EbC8k+/ZBKqAuYi2y/Zm5I32YB0B81JMaA+mA/vW+Oc8g/I7FOdPYaLKR3J5mOf8Ar5+yjcdu8Ek9EFPXaVDVcdWNqKSCVv2ZYx2kfy2+hHctlMc9UxktCGq2RrJGcsgkb3d/TbwKuiqyxOjnhfJFIMruXnHuK13QnQ/bFSxKdxkHbN2+3dBvamq8UrqkWQiO7bUQkY4eMbxuD8irO0PnTnNMUM6R6syMM6+h9sd30OxXM2itR/oHU9fJuLjC12ywH9puz+YD6Fux7vJddYTIMymFqX4jy9pGDt+0N9j8OqC2sZdyNS5ZrHKPFJsbhz1ANx3b+CkmFuZCpM2JuRM2/vEEBo+oIKiHFG/jdRarp0KtiGxjJnfqpA0+w6APYD4dztyAqkqX3VM9Q1BSj9qpN+sjPI5pbsD9fgg6Bpkk10Frt2dO6xJHiOp/kp9itmCJsMLAyNg2aB3BZ+AsyXNPVJpDvI6NocfPcLZICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiIP/9k="

# =========================================================
# REPORT HTML TEMPLATE
# =========================================================
REPORT_HTML = """
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
    .chart-card {
        background: var(--card-bg);
        border-radius: 8px;
        padding: 20px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        margin-bottom: 20px;
        display: flex;
        flex-direction: column;
    }
    .chart-grid {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 20px;
    }
    .chart-container {
        position: relative;
        height: 300px;
        width: 100%;
    }
    @media print {
        @page { size: auto; margin: 10mm; }
        body { background: white !important; color: black !important; }
        .page-wrap { width: 100% !important; margin: 0 !important; padding: 0 !important; }
        .btn, .view-switcher { display: none !important; }
        .chart-grid { display: block !important; }
        .chart-card { break-inside: avoid; box-shadow: none !important; border: 1px solid #eee !important; margin-bottom: 30px !important; page-break-after: auto; }
        .chart-container { height: 350px !important; width: 100% !important; }
        h1, h3 { color: black !important; }
    }
</style>

<div class="page-wrap">
    <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:20px;">
        <h1 style="font-family:'Rajdhani';"><span style="color:var(--accent-gold);">PART</span> ANALYTICS</h1>
        <button class="btn btn-primary" onclick="window.print()">GENERATE PDF</button>
    </div>
    <div class="chart-grid">
        <div class="chart-card">
            <h3>Parts Distribution by Category</h3>
            <div class="chart-container"><canvas id="categoryChart"></canvas></div>
        </div>
        <div class="chart-card">
            <h3>Shipment Status Overview</h3>
            <div class="chart-container"><canvas id="statusChart"></canvas></div>
        </div>
    </div>
    <div class="chart-card">
        <h3>Weekly Shipment Trend</h3>
        <div class="chart-container"><canvas id="trendChart"></canvas></div>
    </div>
</div>

<script>
    const catData = {{ cat_data|safe }};
    const statData = {{ stat_data|safe }};
    const trendData = {{ trend_data|safe }};

    const commonOptions = {
        responsive: true,
        maintainAspectRatio: false,
        plugins: {
            legend: {
                position: 'bottom',
                labels: { boxWidth: 12, color: '#FFFFFF', font: { family: 'Rajdhani', size: 12, weight: 'bold' } }
            }
        }
    };

    const axisStyle = {
        y: { beginAtZero: true, ticks: { color: '#FFFFFF', font: { weight: 'bold', size: 11 } }, grid: { color: 'rgba(255,255,255,0.1)' } },
        x: { ticks: { color: '#FFFFFF', font: { weight: 'bold', size: 11 } }, grid: { display: false } }
    };

    new Chart(document.getElementById('categoryChart'), {
        type: 'doughnut',
        data: {
            labels: catData.map(r => r.category || 'Other'),
            datasets: [{ data: catData.map(r => r.count), backgroundColor: ['#eb0a1e', '#2a6abf', '#c8a84b', '#1a8a5a', '#6a8eaa'], borderColor: 'rgba(255,255,255,0.2)', borderWidth: 2 }]
        },
        options: commonOptions
    });

    new Chart(document.getElementById('statusChart'), {
        type: 'bar',
        data: {
            labels: statData.map(r => r.status),
            datasets: [{ label: 'Count', data: statData.map(r => r.count), backgroundColor: '#3a85e0' }]
        },
        options: { ...commonOptions, plugins: { legend: { display: false } }, scales: axisStyle }
    });

    new Chart(document.getElementById('trendChart'), {
        type: 'line',
        data: {
            labels: trendData.map(r => 'Week ' + r.week),
            datasets: [{ label: 'Total Shipments', data: trendData.map(r => r.count), borderColor: '#c8a84b', backgroundColor: 'rgba(200, 168, 75, 0.1)', fill: true, tension: 0.4, pointBackgroundColor: '#FFFFFF' }]
        },
        options: { ...commonOptions, scales: axisStyle }
    });
</script>
"""

# -------------------------
# Database helpers
# -------------------------
def get_db():
    conn = psycopg2.connect(
        host="localhost",
        database="postgres",
        user="postgres",
        password="Shreya@2005",
        port="5432"
    )
    conn.autocommit = False
    return conn

def dict_cursor(conn):
    return conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

def init_db():
    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute("SELECT current_user, current_database();")
    print(cur.fetchone())
    cur.execute("""
    CREATE TABLE IF NOT EXISTS admins (
        "Sl_No" SERIAL PRIMARY KEY,
        name TEXT,
        username TEXT UNIQUE,
        password TEXT,
        role TEXT DEFAULT 'QACAdmin'
    )
    """)
    cur.execute("SELECT COUNT(*) as cnt FROM admins")
    if cur.fetchone()['cnt'] == 0:
        default_admins = [
            ("QAC Admin", "qacadmin", "qac123", "QACAdmin"),
            ("QIC Admin", "qicadmin", "qic123", "QICAdmin"),
            ("Production Admin", "prodadmin", "prod123", "ProductionAdmin"),
        ]
        cur.executemany('INSERT INTO admins(name,username,password,role) VALUES(%s,%s,%s,%s)', default_admins)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS dealers (
        "Sl_No" SERIAL PRIMARY KEY,
        name TEXT,
        username TEXT UNIQUE,
        password TEXT UNIQUE,
        department TEXT DEFAULT 'General'
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS shipments (
        "Sl_No" SERIAL PRIMARY KEY,
        dealer_id TEXT,
        "Part_Name" TEXT,
        "Part_Number" TEXT,
        "Model" TEXT,
        "Supplier_name" TEXT,
        "Date_sent" TEXT,
        "Customer_Concern" TEXT,
        status TEXT DEFAULT 'Open',
        "Remark" TEXT,
        "PIC" TEXT,
        category TEXT,
        is_deleted INTEGER DEFAULT 0,
        created_by TEXT,
        created_by_role TEXT,
        created_at TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS "Discussion" (
        "Sl_No" SERIAL PRIMARY KEY,
        shipment_id INTEGER,
        pi_number TEXT,
        message TEXT,
        dept TEXT,
        created_at TEXT,
        author_name TEXT,
        author_username TEXT,
        author_role TEXT,
        is_deleted INTEGER DEFAULT 0,
        edited INTEGER DEFAULT 0,
        edited_at TEXT,
        edited_by TEXT
    )
    """)
    try:
        cur.execute("""
            DO $$
            BEGIN
                IF EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='shipments' AND column_name='dealer_id'
                    AND data_type != 'text'
                ) THEN
                    ALTER TABLE shipments ALTER COLUMN dealer_id TYPE TEXT USING dealer_id::TEXT;
                END IF;
            END$$;
        """)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print("dealer_id migration note:", e)

    try:
        cur.execute('ALTER TABLE shipments ADD COLUMN IF NOT EXISTS onedrive_url TEXT')
        conn.commit()
    except Exception:
        conn.rollback()
    conn.commit()
    conn.close()

def get_current_user():
    if 'admin' in session:
        return session.get('admin_name', 'Admin'), session.get('admin_username', 'admin'), session.get('admin_role', 'Admin')
    elif 'dealer' in session:
        return session.get('dealer_name', 'User'), session.get('dealer_username', 'user'), session.get('dealer_department', 'General')
    return 'Unknown', 'unknown', 'Unknown'

def get_base_style():
    return """
<style>
@import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=IBM+Plex+Sans:wght@300;400;500;600;700&display=swap');
:root {
    --steel-darkest: #162d50;
    --steel-darker:  #2d518a;
    --steel-dark:    #2454b4;
    --steel-mid:     #357de9;
    --steel-base:    #2a6abf;
    --steel-light:   #2a6abf;
    --steel-lighter: #3a85e0;
    --steel-pale:    #5ba3f5;
    --steel-frost:   #a8c8f8;
    --steel-ice:     #d4e8ff;
    --chrome-1:      #c8d8e8;
    --chrome-2:      #9ab4cc;
    --chrome-3:      #6a8eaa;
    --accent-gold:   #c8a84b;
    --accent-red:    #c0392b;
    --accent-green:  #1a8a5a;
    --bg-page:       #2c4a83;
    --bg-card:       #0f3986;
    --bg-card2:      #28386e;
    --border-light:  rgba(90,140,200,0.25);
    --border-glow:   rgba(46, 136, 247, 0.5);
    --text-main:     #f5f7fa;
    --text-muted:    #a5c7eb;
    --text-dim:      #a5caf0;
    --shadow-blue:   0 4px 20px rgba(6, 27, 58, 0.4);
    --shadow-glow:   0 0 20px rgba(58,133,224,0.2);
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
    font-family: 'IBM Plex Sans', sans-serif;
    background: var(--bg-page);
    color: var(--text-main);
    min-height: 100vh;
    background-image:
        radial-gradient(ellipse at 20% 20%, rgba(26,74,140,0.15) 0%, transparent 50%),
        radial-gradient(ellipse at 80% 80%, rgba(10,30,70,0.3) 0%, transparent 50%);
}
.site-header {
    background: linear-gradient(135deg, var(--steel-darkest) 0%, var(--steel-dark) 50%, var(--steel-darker) 100%);
    border-bottom: 2px solid transparent;
    border-image: linear-gradient(90deg, transparent, var(--steel-lighter), var(--accent-gold), var(--steel-lighter), transparent) 1;
    padding: 0 24px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    height: 64px;
    position: sticky;
    top: 0;
    z-index: 100;
    box-shadow: 0 2px 30px rgba(0,0,0,0.6);
}
.header-logo-wrap { display: flex; align-items: center; gap: 12px; }
.header-logo-wrap img { height: 44px; filter: drop-shadow(0 0 8px rgba(200,168,75,0.4)); }
.header-brand { display: flex; flex-direction: column; line-height: 1.1; }
.header-brand .brand-main { font-family: 'Rajdhani', sans-serif; font-size: 1.15rem; font-weight: 700; color: var(--steel-ice); letter-spacing: 1px; text-transform: uppercase; }
.header-brand .brand-sub { font-size: 0.68rem; color: var(--accent-gold); letter-spacing: 2px; text-transform: uppercase; font-weight: 500; }
.header-center-title {
    font-family: 'Rajdhani', sans-serif; font-size: 1.5rem; font-weight: 700; letter-spacing: 3px; text-transform: uppercase;
    background: linear-gradient(90deg, var(--steel-frost), var(--steel-ice), var(--accent-gold), var(--steel-ice), var(--steel-frost));
    -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text;
}
.header-right-wrap { display: flex; align-items: center; gap: 10px; }
.user-pill { background: rgba(42,106,191,0.2); border: 1px solid var(--border-light); border-radius: 20px; padding: 4px 14px; font-size: 0.78rem; color: var(--steel-frost); display: flex; align-items: center; gap: 6px; }
.header-nav { display: flex; align-items: center; gap: 2px; }
.header-nav a { color: var(--chrome-2); text-decoration: none; font-size: 0.8rem; font-weight: 500; padding: 5px 10px; border-radius: 4px; transition: all 0.2s; letter-spacing: 0.3px; }
.header-nav a:hover { background: rgba(42,106,191,0.3); color: var(--steel-ice); }
.header-nav a.nav-danger { color: #e88; }
.header-nav a.nav-danger:hover { background: rgba(192,57,43,0.25); color: #ffaaaa; }
.role-tag { display: inline-block; padding: 2px 9px; border-radius: 3px; font-size: 0.7rem; font-weight: 700; letter-spacing: 0.8px; text-transform: uppercase; font-family: 'Rajdhani', sans-serif; }
.role-QACAdmin    { background: linear-gradient(135deg,#1a3a6b,#2a5a9b); color: #a8d8ff; border: 1px solid #2a6abf; }
.role-QICAdmin    { background: linear-gradient(135deg,#0e3a5a,#1a5a7a); color: #a8e8ff; border: 1px solid #1a8ab0; }
.role-ProductionAdmin { background: linear-gradient(135deg,#2a1a5a,#4a2a8a); color: #d8c8ff; border: 1px solid #6a3abf; }
.role-General,.role-dealer,.role-Dealer,.role-QAC,.role-QIC,.role-Production,.role-Supplier { background: linear-gradient(135deg,#1a2a3a,#2a3a4a); color: #8ab8d8; border: 1px solid #3a5a7a; }
.page-wrap { max-width: 1500px; margin: auto; padding: 20px 24px; }
.card { background: linear-gradient(145deg, var(--bg-card) 0%, var(--bg-card2) 100%); border: 1px solid var(--border-light); border-radius: 8px; padding: 20px 22px; margin-bottom: 18px; box-shadow: var(--shadow-blue); position: relative; overflow: hidden; }
.card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px; background: linear-gradient(90deg, transparent, var(--steel-lighter), transparent); }
h1, h2, h3 { font-family: 'Rajdhani', sans-serif; letter-spacing: 0.5px; }
h1 { font-size: 1.6rem; color: var(--steel-ice); }
h2 { font-size: 1.3rem; color: var(--steel-frost); }
h3 { font-size: 1.1rem; color: var(--chrome-2); font-weight: 600; }
.form-row { display: flex; flex-wrap: wrap; gap: 10px; align-items: flex-end; }
.form-group { display: flex; flex-direction: column; gap: 4px; }
.form-group label { font-size: 0.75rem; font-weight: 600; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.5px; }
input[type=text],input[type=password],input[type=date],textarea,select { padding: 7px 11px; background: rgba(10,22,40,0.8); border: 1px solid var(--border-light); border-radius: 5px; color: var(--text-main); font-family: 'IBM Plex Sans', sans-serif; font-size: 0.88rem; transition: border 0.2s, box-shadow 0.2s; }
input:focus,textarea:focus,select:focus { outline: none; border-color: var(--steel-lighter); box-shadow: 0 0 0 3px rgba(58,133,224,0.15); }
select option { background: var(--steel-darkest); color: var(--text-main); }
.btn { display: inline-block; padding: 7px 16px; border: none; border-radius: 5px; font-family: 'Rajdhani', sans-serif; font-size: 0.92rem; font-weight: 600; letter-spacing: 0.5px; cursor: pointer; text-decoration: none; transition: all 0.15s; }
.btn-primary { background: linear-gradient(135deg, var(--steel-mid), var(--steel-base)); color: var(--steel-ice); border: 1px solid var(--steel-light); box-shadow: 0 2px 10px rgba(26,74,140,0.4); }
.btn-primary:hover { background: linear-gradient(135deg, var(--steel-base), var(--steel-lighter)); }
.btn-success { background: linear-gradient(135deg,#0d5a3a,#1a8a5a); color: #a8ffd8; border: 1px solid #1a8a5a; }
.btn-danger  { background: linear-gradient(135deg,#5a0d0d,#8a1a1a); color: #ffaaaa; border: 1px solid #c0392b; }
.btn-warn    { background: linear-gradient(135deg,#5a3a0d,#8a5a1a); color: #ffd8a8; border: 1px solid #b06820; }
.btn-sm { padding: 3px 9px; font-size: 0.78rem; }
.table-wrap { overflow-x: auto; border-radius: 6px; }
table { width: 100%; border-collapse: collapse; font-size: 0.83rem; }
thead tr { background: linear-gradient(135deg, var(--steel-darkest), var(--steel-dark)); border-bottom: 2px solid var(--steel-light); }
th { padding: 10px 11px; text-align: left; font-family: 'Rajdhani', sans-serif; font-weight: 700; font-size: 0.82rem; letter-spacing: 1px; text-transform: uppercase; color: var(--chrome-1); white-space: nowrap; }
td { padding: 9px 11px; border-bottom: 1px solid rgba(42,106,191,0.1); color: var(--text-main); vertical-align: middle; }
tr:hover td { background: rgba(42,106,191,0.07); }
.status-Open            { color: #5af0a0; font-weight: 700; font-size: 0.82rem; }
.status-Inprogress      { color: #f0c050; font-weight: 700; font-size: 0.82rem; }
.status-InprogressTSD   { color: #ffa040; font-weight: 700; font-size: 0.82rem; }
.status-Closed          { color: #f07070; font-weight: 700; font-size: 0.82rem; }
.pagination { display: flex; gap: 5px; margin-top: 14px; flex-wrap: wrap; }
.pagination a { padding: 5px 12px; border: 1px solid var(--border-light); border-radius: 4px; text-decoration: none; font-size: 0.82rem; background: rgba(10,22,40,0.6); color: var(--steel-frost); transition: all 0.15s; }
.pagination a:hover { background: rgba(42,106,191,0.25); border-color: var(--steel-light); }
.pagination a.active { background: linear-gradient(135deg,var(--steel-mid),var(--steel-base)); color: white; border-color: var(--steel-lighter); }
.login-wrap { min-height: 100vh; display: flex; align-items: center; justify-content: center; flex-direction: column; gap: 0; background: radial-gradient(ellipse at 30% 30%, rgba(26,74,140,0.25) 0%, transparent 60%), radial-gradient(ellipse at 70% 70%, rgba(10,30,60,0.4) 0%, transparent 60%), linear-gradient(160deg, #080f1e 0%, #0d1a30 50%, #0a1222 100%); }
.login-header-bar { width: 100%; background: linear-gradient(135deg, var(--steel-darkest), var(--steel-dark)); border-bottom: 2px solid transparent; border-image: linear-gradient(90deg, transparent, var(--steel-lighter), var(--accent-gold), var(--steel-lighter), transparent) 1; padding: 12px 32px; display: flex; align-items: center; justify-content: space-between; box-shadow: 0 2px 30px rgba(0,0,0,0.6); position: fixed; top: 0; z-index: 10; }
.login-content { margin-top: 70px; display: flex; align-items: center; justify-content: center; min-height: calc(100vh - 70px); width: 100%; }
.login-card { background: linear-gradient(145deg, rgba(13,31,60,0.95), rgba(18,28,50,0.98)); border: 1px solid var(--border-light); border-radius: 12px; padding: 36px 32px; width: 100%; max-width: 420px; box-shadow: 0 20px 60px rgba(0,0,0,0.6), 0 0 40px rgba(26,74,140,0.2); position: relative; overflow: hidden; }
.login-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px; background: linear-gradient(90deg, var(--steel-mid), var(--steel-lighter), var(--accent-gold), var(--steel-lighter), var(--steel-mid)); }
.login-logo-area { text-align: center; margin-bottom: 20px; }
.login-logo-area img { height: 72px; filter: drop-shadow(0 0 14px rgba(200,168,75,0.5)); }
.login-title { text-align: center; font-family: 'Rajdhani', sans-serif; font-size: 1.5rem; font-weight: 700; letter-spacing: 2px; text-transform: uppercase; color: var(--steel-ice); margin-bottom: 4px; }
.login-subtitle { text-align: center; color: var(--text-muted); font-size: 0.78rem; letter-spacing: 1px; text-transform: uppercase; margin-bottom: 24px; }
.divider { height: 1px; background: linear-gradient(90deg, transparent, var(--border-light), transparent); margin: 14px 0; }
.field-group { margin-bottom: 14px; }
.field-group label { display: block; font-size: 0.72rem; font-weight: 700; color: var(--text-muted); text-transform: uppercase; letter-spacing: 1px; margin-bottom: 5px; }
.field-group input, .field-group select { width: 100%; }
.pw-wrap { position: relative; }
.pw-wrap input { width: 100%; padding-right: 52px; }
.pw-toggle { position: absolute; right: 10px; top: 50%; transform: translateY(-50%); background: none; border: none; cursor: pointer; color: var(--text-muted); font-size: 0.75rem; font-family: inherit; padding: 2px 6px; border-radius: 3px; }
.pw-toggle:hover { color: var(--steel-frost); background: rgba(42,106,191,0.2); }
.login-btn { width: 100%; padding: 10px; margin-top: 8px; font-family: 'Rajdhani', sans-serif; font-size: 1rem; font-weight: 700; letter-spacing: 2px; text-transform: uppercase; background: linear-gradient(135deg, var(--steel-mid) 0%, var(--steel-base) 50%, var(--steel-light) 100%); color: var(--steel-ice); border: 1px solid var(--steel-lighter); border-radius: 6px; cursor: pointer; transition: all 0.2s; box-shadow: 0 3px 15px rgba(26,74,140,0.4); }
.login-btn:hover { background: linear-gradient(135deg, var(--steel-base), var(--steel-lighter)); box-shadow: 0 5px 20px rgba(42,106,191,0.5); transform: translateY(-1px); }
.login-links { text-align: center; margin-top: 16px; font-size: 0.82rem; color: var(--text-muted); }
.login-links a { color: var(--steel-pale); text-decoration: none; }
.login-links a:hover { color: var(--steel-ice); text-decoration: underline; }
.alert { padding: 10px 14px; border-radius: 6px; margin-bottom: 14px; font-size: 0.88rem; }
.alert-info    { background: rgba(26,74,140,0.2); border-left: 3px solid var(--steel-lighter); color: var(--steel-frost); }
.alert-success { background: rgba(26,138,90,0.15); border-left: 3px solid #1a8a5a; color: #5af0a0; }
.alert-danger  { background: rgba(192,57,43,0.15); border-left: 3px solid #c0392b; color: #ffaaaa; }
.chat-container { display: flex; flex-direction: column; gap: 10px; padding: 10px 4px; max-height: 600px; overflow-y: auto; }
.chat-bubble-wrap { display: flex; flex-direction: column; max-width: 68%; }
.chat-bubble-wrap.left  { align-self: flex-start; align-items: flex-start; }
.chat-bubble-wrap.right { align-self: flex-end;   align-items: flex-end; }
.chat-bubble { position: relative; padding: 10px 14px 8px 14px; border-radius: 14px; font-size: 0.88rem; line-height: 1.5; word-break: break-word; box-shadow: 0 3px 12px rgba(0,0,0,0.35); }
.chat-bubble.left { background: linear-gradient(145deg, #0e2a50, #1a4080); border: 1px solid rgba(58,133,224,0.35); border-top-left-radius: 4px; color: var(--text-main); }
.chat-bubble.left::before { content: ''; position: absolute; left: -8px; top: 12px; border-width: 6px 9px 6px 0; border-style: solid; border-color: transparent #1a4080 transparent transparent; }
.chat-bubble.right { background: linear-gradient(145deg, #0a3a2a, #155a3a); border: 1px solid rgba(26,138,90,0.4); border-top-right-radius: 4px; color: #d0ffe8; }
.chat-bubble.right::before { content: ''; position: absolute; right: -8px; top: 12px; border-width: 6px 0 6px 9px; border-style: solid; border-color: transparent transparent transparent #155a3a; }
.chat-bubble.deleted-bubble { background: rgba(10,14,20,0.5) !important; border: 1px dashed rgba(120,120,140,0.3) !important; color: var(--text-dim) !important; font-style: italic; }
.chat-bubble.deleted-bubble::before { display: none; }
.bubble-header { display: flex; align-items: center; gap: 7px; margin-bottom: 5px; flex-wrap: wrap; }
.bubble-author { font-family: 'Rajdhani', sans-serif; font-weight: 700; font-size: 0.88rem; letter-spacing: 0.3px; }
.left  .bubble-author { color: var(--steel-pale); }
.right .bubble-author { color: #5af0a0; }
.bubble-pi { font-size: 0.72rem; background: rgba(200,168,75,0.18); border: 1px solid rgba(200,168,75,0.35); color: var(--accent-gold); padding: 1px 6px; border-radius: 3px; font-family: 'Rajdhani', sans-serif; font-weight: 600; letter-spacing: 0.5px; }
.bubble-time { font-size: 0.69rem; color: rgba(180,200,230,0.55); margin-top: 5px; text-align: right; display: flex; align-items: center; justify-content: flex-end; gap: 5px; }
.left .bubble-time { justify-content: flex-start; }
.edited-tag { font-size: 0.69rem; color: var(--accent-gold); font-style: italic; }
.bubble-actions { display: flex; gap: 5px; margin-top: 5px; }
.bubble-edit-form { margin-top: 8px; background: rgba(0,0,0,0.2); border-radius: 8px; padding: 8px; }
.chat-bubble::after { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 40%; border-radius: inherit; background: linear-gradient(180deg, rgba(255,255,255,0.06) 0%, transparent 100%); pointer-events: none; }
.onedrive-bar { background: linear-gradient(145deg, rgba(10,30,60,0.9), rgba(20,40,80,0.95)); border: 1px solid rgba(200,168,75,0.35); border-radius: 8px; padding: 12px 16px; margin-bottom: 14px; display: flex; align-items: center; gap: 12px; flex-wrap: wrap; }
.onedrive-bar label { font-size: 0.75rem; font-weight: 700; color: var(--accent-gold); text-transform: uppercase; letter-spacing: 1px; white-space: nowrap; }
.onedrive-bar input[type=text] { flex: 1; min-width: 260px; }
.onedrive-bar a.btn { white-space: nowrap; }
::-webkit-scrollbar { width: 7px; height: 7px; }
::-webkit-scrollbar-track { background: var(--steel-darkest); }
::-webkit-scrollbar-thumb { background: var(--steel-mid); border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: var(--steel-light); }
code { font-family: monospace; color: var(--steel-pale); font-size: 0.85em; }
.view-switcher {
    position: fixed; bottom: 0; left: 0; right: 0;
    display: flex; z-index: 200;
    box-shadow: 0 -3px 20px rgba(0,0,0,0.5);
}
.view-switcher a {
    flex: 1; display: flex; align-items: center; justify-content: center; gap: 10px;
    padding: 13px 0; font-family: 'Rajdhani', sans-serif; font-size: 1.05rem;
    font-weight: 700; letter-spacing: 2px; text-transform: uppercase;
    text-decoration: none; transition: all 0.18s; border-top: 2px solid transparent;
}
.view-switcher a.vs-normal { background: linear-gradient(135deg, var(--steel-darkest) 0%, var(--steel-dark) 100%); color: var(--steel-frost); border-color: var(--steel-lighter); border-right: 1px solid rgba(90,140,200,0.3); }
.view-switcher a.vs-normal:hover, .view-switcher a.vs-normal.vs-active { background: linear-gradient(135deg, var(--steel-dark) 0%, var(--steel-mid) 100%); color: var(--steel-ice); border-color: var(--accent-gold); }
.view-switcher a.vs-summary { background: linear-gradient(135deg, #0d3a22 0%, #155a30 100%); color: #7adfa8; border-color: #1a8a5a; }
.view-switcher a.vs-summary:hover, .view-switcher a.vs-summary.vs-active { background: linear-gradient(135deg, #155a30 0%, #1da85a 100%); color: #d0ffe8; border-color: #2adf7a; }
.view-switcher a .vs-icon { font-size: 1.2rem; }
.page-wrap { padding-bottom: 60px; }
</style>
<script>
function togglePW(inputId, btnId) {
    const inp = document.getElementById(inputId);
    const btn = document.getElementById(btnId);
    if (inp.type === 'password') { inp.type = 'text'; btn.textContent = 'Hide'; }
    else { inp.type = 'password'; btn.textContent = 'Show'; }
}
</script>
"""


def get_status_css_class(status):
    if not status:
        return 'status-Open'
    clean = status.replace(' ', '').replace(':', '')
    return f'status-{clean}'


# =========================================================
# HEADERS
# =========================================================
def render_admin_header():
    name, username, role = get_current_user()
    is_admin = 'admin' in session
    role_cls = ''.join(c for c in role if c.isalnum())
    if is_admin:
        nav = """
        <a href="/admin_dashboard">Dashboard</a>
        <a href="/reports">Reports</a>
        <a href="/add_shipment">+ New Part</a>
        <a href="/manage_admins">Admins</a>
        <a href="/manage_users">Users</a>
        <a href="/trash">Trash</a>
        <a href="/logout" class="nav-danger">Logout</a>
        """
    elif 'dealer' in session:
        nav = '<a href="/dealer_dashboard">Dashboard</a><a href="/reports">Reports</a><a href="/add_shipment">+ New Part</a><a href="/trash">Trash</a><a href="/logout" class="nav-danger">Logout</a>'
    else:
        nav = '<a href="/">Login</a>'

    return f"""
    <div class="site-header">
      <div class="header-logo-wrap">
        <img src="{TOYOTA_LOGO}" alt="Toyota" style="height:44px;filter:drop-shadow(0 0 8px rgba(200,168,75,0.4))">
        <div class="header-brand">
          <span class="brand-main">TKM</span>
          <span class="brand-sub">Toyota Kirloskar Motors</span>
        </div>
      </div>
      <div class="header-center-title">QAC Discussion Portal</div>
      <div class="header-right-wrap">
        <span class="role-tag role-{role_cls}">{role}</span>
        <span class="user-pill">&#128100; {name} &bull; {username}</span>
        <div class="header-nav">{nav}</div>
      </div>
    </div>
    """

def render_dealer_header():
    name, username, role = get_current_user()
    is_admin = 'admin' in session
    role_cls = ''.join(c for c in role if c.isalnum())
    if is_admin:
        nav = """
        <a href="/admin_dashboard">Dashboard</a>
        <a href="/reports">Reports</a>
        <a href="/add_shipment">+ New Part</a>
        <a href="/trash">Trash</a>
        <a href="/logout" class="nav-danger">Logout</a>
        """
    elif 'dealer' in session:
        nav = '<a href="/dealer_dashboard">Dashboard</a><a href="/reports">Reports</a><a href="/add_shipment">+ New Part</a><a href="/trash">Trash</a><a href="/logout" class="nav-danger">Logout</a>'
    else:
        nav = '<a href="/">Login</a>'

    return f"""
    <div class="site-header">
      <div class="header-logo-wrap">
        <img src="{TOYOTA_LOGO}" alt="Toyota" style="height:44px;filter:drop-shadow(0 0 8px rgba(200,168,75,0.4))">
        <div class="header-brand">
          <span class="brand-main">TKM</span>
          <span class="brand-sub">Toyota Kirloskar Motors</span>
        </div>
      </div>
      <div class="header-center-title">QAC Discussion Portal</div>
      <div class="header-right-wrap">
        <span class="role-tag role-{role_cls}">{role}</span>
        <span class="user-pill">&#128100; {name} &bull; {username}</span>
        <div class="header-nav">{nav}</div>
      </div>
    </div>
    """

def render_login_header():
    return f"""
    <div class="login-header-bar">
      <div class="header-logo-wrap">
        <img src="{TOYOTA_LOGO}" alt="Toyota" style="height:40px;filter:drop-shadow(0 0 6px rgba(200,168,75,0.4))">
        <div class="header-brand">
          <span class="brand-main">TKM</span>
          <span class="brand-sub">Toyota Kirloskar Motors</span>
        </div>
      </div>
      <div class="header-center-title" style="font-size:1.1rem;">QAC Discussion Portal</div>
      <div style="font-size:0.72rem;color:var(--text-muted);letter-spacing:1px;text-transform:uppercase;">NTF Discussion System</div>
    </div>
    """


# =========================================================
# AUTH ROUTES
# =========================================================
@app.route('/')
def home():
    html = get_base_style() + render_login_header() + f"""
    <div class="login-content">
    <div class="login-card">
      <div class="login-logo-area"><img src="{TOYOTA_LOGO}" alt="Toyota Logo"></div>
      <div class="login-title">User Login</div>
      <div class="login-subtitle">QAC Discussion Portal</div>
      <div class="divider"></div>
      <form method="POST" action="/login">
        <div class="field-group"><label>Username</label>
          <input type="text" name="username" required placeholder="Enter your username"></div>
        <div class="field-group"><label>Password</label>
          <div class="pw-wrap">
            <input type="password" name="password" id="lp" required placeholder="Enter your password">
            <button type="button" class="pw-toggle" id="lpt" onclick="togglePW('lp','lpt')">Show</button>
          </div></div>
        <button type="submit" class="login-btn">Login</button>
      </form>
      <div class="login-links">
        <a href="/register">Register as User</a> &nbsp;&nbsp;|&nbsp;&nbsp; <a href="/admin_login">Admin Login</a>
      </div>
    </div>
    </div>
    """
    return render_template_string(html)


@app.route('/register', methods=['GET', 'POST'])
def register():
    error_msg = ""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        name     = request.form.get('name', '').strip()
        password = request.form.get('password', '').strip()
        dept     = request.form.get('department', 'General')

        if not username or not name or not password:
            error_msg = "All fields are required."
        else:
            conn = get_db()
            cur = dict_cursor(conn)
            cur.execute("SELECT COUNT(*) as cnt FROM dealers WHERE username=%s", (username,))
            if cur.fetchone()['cnt'] > 0:
                conn.close()
                error_msg = f"Username <strong>{username}</strong> is already taken. Please choose a different username."
            else:
                try:
                    cur.execute(
                        "INSERT INTO dealers(name,username,password,department) VALUES(%s,%s,%s,%s)",
                        (name, username, password, dept)
                    )
                    conn.commit()
                    conn.close()
                    return redirect('/')
                except Exception as e:
                    conn.rollback()
                    conn.close()
                    error_msg = "Registration failed. Please try again."

    alert_html = f'<div class="alert alert-danger">{error_msg}</div>' if error_msg else ''
    html = get_base_style() + render_login_header() + f"""
    <div class="login-content">
    <div class="login-card">
      <div class="login-logo-area"><img src="{TOYOTA_LOGO}" alt="Toyota Logo"></div>
      <div class="login-title">Register</div>
      <div class="login-subtitle">Create your user account</div>
      <div class="divider"></div>
      {alert_html}
      <form method="POST" action="/register">
        <div class="field-group"><label>Full Name</label>
          <input type="text" name="name" required placeholder="Your full name" value="{request.form.get('name','') if request.method=='POST' else ''}"></div>
        <div class="field-group"><label>Username</label>
          <input type="text" name="username" required placeholder="Choose a username" value="{request.form.get('username','') if request.method=='POST' else ''}"></div>
        <div class="field-group"><label>Department</label>
          <select name="department">
            <option value="General">TSD</option>
            <option value="QAC">QAC</option>
            <option value="QIC">QIC</option>
            <option value="Production">Production</option>
            <option value="Dealer">Dealer</option>
            <option value="Supplier">Supplier</option>
          </select></div>
        <div class="field-group"><label>Password</label>
          <div class="pw-wrap">
            <input type="password" name="password" id="rp" required placeholder="Choose a password">
            <button type="button" class="pw-toggle" id="rpt" onclick="togglePW('rp','rpt')">Show</button>
          </div></div>
        <button type="submit" class="login-btn">Register</button>
      </form>
      <div class="login-links"><a href="/">Back to Login</a></div>
    </div></div>
    """
    return render_template_string(html)


@app.route('/login', methods=['POST'])
def login():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute("SELECT * FROM dealers WHERE username=%s AND password=%s", (username, password))
    user = cur.fetchone()
    conn.close()
    if user:
        session['dealer'] = user['Sl_No']
        session['dealer_name'] = user['name']
        session['dealer_username'] = user['username']
        session['dealer_department'] = user['department']
        return redirect('/dealer_dashboard')
    return render_template_string(get_base_style() + render_login_header() + f"""
        <div class="login-content"><div class="login-card">
        <div class="login-logo-area"><img src="{TOYOTA_LOGO}" alt="Toyota Logo"></div>
        <div class="login-title">User Login</div>
        <div class="login-subtitle">QAC Discussion Portal</div>
        <div class="divider"></div>
        <div class="alert alert-danger">Invalid username or password. <a href="/" style="color:var(--steel-pale)">Try again</a></div>
        </div></div>""")


@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    error_msg = ""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        conn = get_db()
        cur = dict_cursor(conn)
        cur.execute("SELECT * FROM admins WHERE username=%s AND password=%s", (username, password))
        admin = cur.fetchone()
        conn.close()
        if admin:
            session['admin'] = admin['Sl_No']
            session['admin_name'] = admin['name']
            session['admin_username'] = admin['username']
            session['admin_role'] = admin['role']
            return redirect('/admin_dashboard')
        error_msg = "Invalid admin credentials."

    alert_html = f'<div class="alert alert-danger">{error_msg}</div>' if error_msg else ''
    html = get_base_style() + render_login_header() + f"""
    <div class="login-content">
    <div class="login-card">
      <div class="login-logo-area"><img src="{TOYOTA_LOGO}" alt="Toyota Logo"></div>
      <div class="login-title">Admin Login</div>
      <div class="login-subtitle">QAC &bull; QIC &bull; Production</div>
      <div class="divider"></div>
      {alert_html}
      <form method="POST" action="/admin_login">
        <div class="field-group"><label>Username</label>
          <input type="text" name="username" required placeholder="Admin username"></div>
        <div class="field-group"><label>Password</label>
          <div class="pw-wrap">
            <input type="password" name="password" id="ap" required placeholder="Admin password">
            <button type="button" class="pw-toggle" id="apt" onclick="togglePW('ap','apt')">Show</button>
          </div></div>
        <button type="submit" class="login-btn">Login as Admin</button>
      </form>
      <div class="login-links"><a href="/">User Login</a></div>
    </div></div>
    """
    return render_template_string(html)


@app.route('/export_discussion/<int:shipment_id>')
def export_discussion(shipment_id):
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')
    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute("""
        SELECT pi_number, message, dept, created_at,
               author_name, author_username, author_role,
               edited, edited_at, edited_by
        FROM "Discussion"
        WHERE shipment_id=%s AND is_deleted=0
        ORDER BY "Sl_No" ASC
    """, (shipment_id,))
    rows = cur.fetchall()
    conn.close()
    data = [dict(r) for r in rows]
    df = pd.DataFrame(data)
    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    return send_file(output, download_name=f"discussion_{shipment_id}.xlsx", as_attachment=True)


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')


# =========================================================
# DASHBOARD HELPERS
# =========================================================
def build_shipment_table(shipments, is_admin):
    rows = ""
    for idx, s in enumerate(shipments, start=1):
        current_status = s.get('status') or 'Open'
        status_cls = get_status_css_class(current_status)
        dealer_id_val = s.get('dealer_id') or ''

        actions = f"""
        <a href="{url_for('edit_shipment', slno=s['Sl_No'])}" class="btn btn-sm btn-warn">Edit</a>
        <a href="{url_for('delete_shipment', slno=s['Sl_No'])}" class="btn btn-sm btn-danger"
           onclick="return confirm('Move to trash?')">Delete</a>
        """

        rows += f"""
        <tr>
          <td style="color:var(--text-muted); font-weight: bold;">{idx}</td>
          <td>{s['Date_sent'] or ''}</td>
          <td>{s['Model'] or ''}</td>
          <td><code>{s['Part_Number'] or ''}</code></td>
          <td><strong style="color:var(--steel-frost)">{s['Part_Name'] or ''}</strong></td>
          <td>{s['Supplier_name'] or ''}</td>
          <td style="max-width:160px;white-space:normal;font-size:.8rem">{s['Customer_Concern'] or ''}</td>
          <td>{s['PIC'] or ''}</td>
          <td>{s['category'] or ''}</td>
          <td>{s['Remark'] or ''}</td>
          <td><span class="{status_cls}">{current_status}</span></td>
          <td style="color:var(--accent-gold);font-weight:600">{dealer_id_val}</td>
          <td><a href="{url_for('view_discussion', shipment_id=s['Sl_No'])}" class="btn btn-sm btn-success">&#128172; Discuss</a></td>
          <td>{actions}</td>
        </tr>"""
    return rows

def build_filter_form(vals, action):
    def sel(name, field, options):
        opts = '<option value="">All</option>'
        for v in options:
            sel_attr = 'selected' if vals.get(field) == v else ''
            opts += f'<option value="{v}" {sel_attr}>{v}</option>'
        return f'<select name="{field}">{opts}</select>'
    return f"""
    <form method="get" action="{action}" class="form-row">
      <div class="form-group"><label>Search</label>
        <input type="text" name="query" value="{vals.get('query','')}" placeholder="Part Name / Number" style="min-width:160px"></div>
      <div class="form-group"><label>Model</label>
        <input type="text" name="model" value="{vals.get('model','')}" placeholder="Model"></div>
      <div class="form-group"><label>Supplier</label>
        <input type="text" name="supplier" value="{vals.get('supplier','')}" placeholder="Supplier"></div>
      <div class="form-group"><label>Date</label>
        <input type="date" name="date_sent" value="{vals.get('date_sent','')}"></div>
      <div class="form-group"><label>PI Number</label>
        <input type="text" name="pic" value="{vals.get('pic','')}" placeholder="PI Number"></div>
      <div class="form-group"><label>Status</label>
        {sel('Status','status',['Open','Inprogress','Inprogress : TSD','Closed'])}</div>
      <div class="form-group"><label>Remark</label>
        {sel('Remark','remark',['external','NTF','misjudgement'])}</div>
      <div class="form-group"><label>Category</label>
        {sel('Category','category',['electrical','body','chassis','engine'])}</div>
      <div class="form-group"><label>&nbsp;</label>
        <button type="submit" class="btn btn-primary">&#128269; Filter</button></div>
    </form>"""


# =========================================================
# DASHBOARDS
# =========================================================
def _dashboard_query(vals, page):
    sql_base = 'FROM shipments WHERE is_deleted=0'
    params = []
    if vals['query']:
        sql_base += ' AND ("Part_Name" ILIKE %s OR "Part_Number" ILIKE %s)'
        params.extend([f"%{vals['query']}%"] * 2)
    if vals['model']:
        sql_base += ' AND "Model" ILIKE %s'; params.append(f"%{vals['model']}%")
    if vals['supplier']:
        sql_base += ' AND "Supplier_name" ILIKE %s'; params.append(f"%{vals['supplier']}%")
    if vals['date_sent']:
        sql_base += ' AND "Date_sent"=%s'; params.append(vals['date_sent'])
    if vals['pic']:
        sql_base += ' AND "PIC" ILIKE %s'; params.append(f"%{vals['pic']}%")
    if vals['status']:
        sql_base += ' AND status=%s'; params.append(vals['status'])
    if vals['remark']:
        sql_base += ' AND "Remark"=%s'; params.append(vals['remark'])
    if vals['category']:
        sql_base += ' AND category=%s'; params.append(vals['category'])
    offset = (page - 1) * PER_PAGE
    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute(f'SELECT COUNT(*) as cnt {sql_base}', params)
    total = cur.fetchone()['cnt']
    cur.execute(f'SELECT * {sql_base} ORDER BY "Sl_No" DESC LIMIT %s OFFSET %s', params + [PER_PAGE, offset])
    shipments = cur.fetchall()
    conn.close()
    return shipments, total


@app.route('/admin_dashboard')
def admin_dashboard():
    if 'admin' not in session:
        return redirect('/admin_login')
    vals = {k: request.args.get(k, '') for k in ['query','model','supplier','date_sent','pic','status','remark','category']}
    page = int(request.args.get('page', 1))
    shipments, total = _dashboard_query(vals, page)
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    rows = build_shipment_table(shipments, True)
    qs = '&'.join(f"{k}={v}" for k, v in vals.items() if v)
    pagination = "".join([f'<a href="?page={p}&{qs}" class="{"active" if p==page else ""}">{p}</a>' for p in range(1, total_pages+1)])
    html = get_base_style() + render_admin_header() + f"""
    <div class="page-wrap">
      <div class="card">{build_filter_form(vals, '/admin_dashboard')}</div>
      <div class="card">
        <h3 style="margin-bottom:12px">&#128230; Parts / Shipments &nbsp;
          <small style="color:var(--text-dim);font-weight:400;font-family:'IBM Plex Sans',sans-serif;font-size:.85rem">{total} records</small></h3>
        <div class="table-wrap"><table>
          <thead><tr><th>SL</th><th>Date</th><th>Model</th><th>Part No.</th><th>Part Name</th>
            <th>Supplier</th><th>Concern</th><th>PI Number</th><th>Category</th>
            <th>Remark</th><th>Status</th><th>Dealer ID</th><th>Discussion</th><th>Actions</th></tr></thead>
          <tbody>{rows}</tbody>
        </table></div>
        <div class="pagination">{pagination}</div>
      </div>
    </div>
    <div class="view-switcher">
      <a href="/admin_dashboard" class="vs-normal vs-active"><span class="vs-icon">&#9776;</span> Normal View</a>
      <a href="/summary_dashboard" class="vs-summary"><span class="vs-icon">&#128202;</span> Summary View</a>
    </div>
    """
    return render_template_string(html)


@app.route('/dealer_dashboard')
def dealer_dashboard():
    if 'dealer' not in session:
        return redirect('/')

    vals = {k: request.args.get(k, '') for k in ['query','model','supplier','date_sent','pic','status','remark','category']}
    page = int(request.args.get('page', 1))
    shipments, total = _dashboard_query(vals, page)
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    rows = build_shipment_table(shipments, True)
    qs = '&'.join(f"{k}={v}" for k, v in vals.items() if v)
    pagination = "".join([f'<a href="?page={p}&{qs}" class="{"active" if p==page else ""}">{p}</a>' for p in range(1, total_pages+1)])

    html = get_base_style() + render_dealer_header() + f"""
    <div class="page-wrap">
      <div class="card">{build_filter_form(vals, '/dealer_dashboard')}</div>
      <div class="card">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px;">
            <h3 style="margin: 0;">
                &#128230; Available Parts &nbsp;
                <small style="color:var(--text-dim); font-weight:400; font-family:'IBM Plex Sans',sans-serif">({total} records)</small>
            </h3>
            <div style="display: flex; gap: 10px;">
                <a href="/add_shipment" class="btn btn-sm btn-primary" style="text-decoration:none">+ New Part</a>
                <a href="/trash" class="btn btn-sm" style="background:rgba(220, 53, 69, 0.1); color:#dc3545; border:1px solid #dc3545; text-decoration:none;">
                    &#128465; Trash
                </a>
            </div>
        </div>
        <div class="table-wrap">
            <table>
              <thead>
                <tr>
                  <th>SL</th><th>Date</th><th>Model</th><th>Part No.</th><th>Part Name</th>
                  <th>Supplier</th><th>Concern</th><th>PI Number</th><th>Category</th>
                  <th>Remark</th><th>Status</th><th>Dealer ID</th><th>Discussion</th><th>Actions</th>
                </tr>
              </thead>
              <tbody>{rows}</tbody>
            </table>
        </div>
        <div class="pagination">{pagination}</div>
      </div>
    </div>
    """
    return render_template_string(html)


# =========================================================
# ADD / EDIT SHIPMENT
# =========================================================
def shipment_form(action, data=None, btn="Add Part"):
    d = data or {}
    def v(k):
        if not d: return ''
        try:
            val = d.get(k, '') if isinstance(d, dict) else (d[k] if k in d.keys() else '')
            return val if val is not None else ''
        except: return ''

    def sel(name, opts, cur):
        o = "".join([f'<option value="{x}" {"selected" if str(cur)==str(x) else ""}>{x if x else "--Select--"}</option>' for x in opts])
        return f'<select name="{name}">{o}</select>'

    status_options = ['Open', 'Inprogress', 'Inprogress : TSD', 'Closed']

    return f"""
    <form method="POST" action="{action}">
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;">
      <div class="form-group"><label>Dealer ID *</label>
        <input type="text" name="dealer_id" value="{v('dealer_id')}" required placeholder="e.g. MU01A"></div>
      <div class="form-group"><label>Part Name *</label>
        <input type="text" name="part_name" value="{v('Part_Name')}" required></div>
      <div class="form-group"><label>Part Number *</label>
        <input type="text" name="part_number" value="{v('Part_Number')}" required></div>
      <div class="form-group"><label>Model</label>
        <input type="text" name="model" value="{v('Model')}"></div>
      <div class="form-group"><label>Supplier Name</label>
        <input type="text" name="supplier" value="{v('Supplier_name')}"></div>
      <div class="form-group"><label>Date</label>
        <input type="date" name="date_sent" value="{v('Date_sent')}"></div>
      <div class="form-group"><label>PI Number</label>
        <input type="text" name="pic" value="{v('PIC')}"></div>
      <div class="form-group"><label>Status</label>
        {sel('status', status_options, v('status') or 'Open')}</div>
      <div class="form-group"><label>Remark</label>
        {sel('remark', ['','external','NTF','misjudgement'], v('Remark'))}</div>
      <div class="form-group"><label>Category</label>
        {sel('category', ['','electrical','body','chassis','engine'], v('category'))}</div>
    </div>
    <div class="form-group" style="margin-top:12px"><label>Customer Concern</label>
      <textarea name="customer_concern" rows="3" style="width:100%">{v('Customer_Concern')}</textarea></div>
    <br>
    <button type="submit" class="btn btn-primary">{btn}</button>
    <a href="/admin_dashboard" class="btn" style="background:rgba(42,106,191,0.15);color:var(--chrome-2);border:1px solid var(--border-light);text-decoration:none;padding:8px 15px;display:inline-block;border-radius:4px;">Cancel</a>
    </form>"""


@app.route('/add_shipment', methods=['GET', 'POST'])
def add_shipment():
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')

    if request.method == 'POST':
        name, username, role = get_current_user()
        conn = get_db()
        cur = dict_cursor(conn)
        cur.execute("""INSERT INTO shipments
        (dealer_id, "Part_Name","Part_Number","Model","Supplier_name","Date_sent",status,"Remark","PIC",category,"Customer_Concern",created_by,created_by_role,created_at)
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        (request.form['dealer_id'], request.form['part_name'], request.form['part_number'], request.form['model'],
         request.form['supplier'], request.form['date_sent'], request.form['status'],
         request.form['remark'], request.form['pic'], request.form['category'],
         request.form['customer_concern'], username, role, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()
        return redirect('/admin_dashboard' if 'admin' in session else '/dealer_dashboard')

    html = get_base_style() + render_admin_header() + f"""
    <div class="page-wrap"><div class="card">
      <h2>&#10133; New Part / Shipment</h2><br>
      {shipment_form('/add_shipment', btn='Add Part')}
    </div></div>"""
    return render_template_string(html)


@app.route('/edit_shipment/<int:slno>', methods=['GET', 'POST'])
def edit_shipment(slno):
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')
    conn = get_db()
    cur = dict_cursor(conn)
    if request.method == 'POST':
        cur.execute("""UPDATE shipments SET
            dealer_id=%s, "Part_Name"=%s, "Part_Number"=%s, "Model"=%s, "Supplier_name"=%s,
            "Date_sent"=%s, status=%s, "Remark"=%s, "PIC"=%s, category=%s, "Customer_Concern"=%s
            WHERE "Sl_No"=%s""",
            (request.form['dealer_id'], request.form['part_name'], request.form['part_number'],
             request.form['model'], request.form['supplier'], request.form['date_sent'],
             request.form['status'], request.form['remark'], request.form['pic'],
             request.form['category'], request.form['customer_concern'], slno))
        conn.commit()
        conn.close()
        return redirect('/admin_dashboard' if 'admin' in session else '/dealer_dashboard')

    cur.execute('SELECT * FROM shipments WHERE "Sl_No"=%s', (slno,))
    shipment = cur.fetchone()
    conn.close()
    return render_template_string(get_base_style() + render_admin_header() + f"""
        <div class="page-wrap"><div class="card">
            <h2>Edit Shipment #{slno}</h2><br>
            {shipment_form(f'/edit_shipment/{slno}', dict(shipment), btn='Update Part')}
        </div></div>""")


# =========================================================
# SUMMARY DASHBOARD
# =========================================================
def get_summary_data(vals):
    conn = get_db()
    cur = dict_cursor(conn)

    sql = """
        SELECT * FROM (
            SELECT "Part_Name", "Model", COUNT(*) as total_cases,
                   MAX(status) as latest_status,
                   STRING_AGG("Sl_No"::TEXT, ',') as shipment_ids
            FROM shipments
            WHERE is_deleted=0
            GROUP BY "Part_Name", "Model"
        ) as summary
        WHERE 1=1
    """
    params = []

    if vals.get('part_name'):
        sql += ' AND "Part_Name" ILIKE %s'
        params.append(f"%{vals['part_name']}%")

    if vals.get('model'):
        sql += ' AND "Model" ILIKE %s'
        params.append(f"%{vals['model']}%")

    if vals.get('status'):
        sql += ' AND latest_status = %s'
        params.append(vals['status'])

    if vals.get('total_cases'):
        sql += ' AND total_cases = %s'
        params.append(vals['total_cases'])

    sql += ' ORDER BY "Part_Name"'

    cur.execute(sql, params)
    data = cur.fetchall()
    conn.close()
    return data


# =========================================================
# ── NEW: Export ALL Summary to Excel ──────────────────────
# =========================================================
@app.route('/export_summary_excel')
def export_summary_excel():
    """
    Export the full summary (Part Name, Model, Total Cases/Quantity, Status)
    to a professionally styled Excel file. Respects the same filters as
    the summary_dashboard so users can also export a filtered subset.
    """
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')

    vals = {
        'part_name':   request.args.get('part_name', ''),
        'model':       request.args.get('model', ''),
        'total_cases': request.args.get('total_cases', ''),
        'status':      request.args.get('status', '')
    }

    data = get_summary_data(vals)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary Report"

    # ── Styles ──────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", start_color="1A3A6B", end_color="1A3A6B")
    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill("solid", start_color="162D50", end_color="162D50")
    title_font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    sub_fill   = PatternFill("solid", start_color="0F3986", end_color="0F3986")
    sub_font   = Font(name="Arial", italic=True, size=9,  color="7EB6FF")
    tot_fill   = PatternFill("solid", start_color="1A3A6B", end_color="1A3A6B")
    tot_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    even_fill  = PatternFill("solid", start_color="D4E8FF", end_color="D4E8FF")
    odd_fill   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    data_font  = Font(name="Arial", size=10)
    bold_font  = Font(name="Arial", size=10, bold=True)

    thin = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Row 1: Title banner ──────────────────────────────────
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = "QAC Discussion Portal — Summary Report"
    c.font      = title_font
    c.fill      = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 32

    # ── Row 2: Subtitle ──────────────────────────────────────
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value     = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                   f"   |   Total Parts: {len(data)}")
    c.font      = sub_font
    c.fill      = sub_fill
    c.alignment = center
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ────────────────────────────────
    headers    = ["Sl. No.", "Part Name", "Model", "Total Cases (Qty)", "Latest Status"]
    col_widths = [10, 38, 22, 22, 24]
    for ci, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = center
        c.border    = thin
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 28

    # ── Data rows ────────────────────────────────────────────
    status_colors = {
        'closed':     "C0392B",
        'tsd':        "B06820",
        'inprogress': "2A6ABF",
        'open':       "1A8A5A",
    }

    for ri, d in enumerate(data, start=4):
        fill = even_fill if ri % 2 == 0 else odd_fill
        row_vals = [
            ri - 3,
            d['Part_Name']    or '',
            d['Model']        or '',
            int(d['total_cases']),
            d['latest_status'] or 'Open',
        ]
        for ci, val in enumerate(row_vals, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill   = fill
            c.border = thin
            c.font   = bold_font if ci == 4 else data_font
            c.alignment = center if ci in (1, 3, 4, 5) else left

        # Colour-code status cell
        sv = (d['latest_status'] or 'Open').lower()
        if 'closed' in sv:
            color = status_colors['closed']
        elif 'tsd' in sv:
            color = status_colors['tsd']
        elif 'inprogress' in sv or 'in progress' in sv:
            color = status_colors['inprogress']
        else:
            color = status_colors['open']
        ws.cell(row=ri, column=5).font = Font(name="Arial", size=10, bold=True, color=color)

    # ── Totals row ───────────────────────────────────────────
    tr = len(data) + 4
    for ci in range(1, 6):
        c = ws.cell(row=tr, column=ci)
        c.fill   = tot_fill
        c.border = thin
        c.font   = tot_font
    ws.cell(row=tr, column=1, value="TOTAL").alignment = center
    ws.cell(row=tr, column=2, value=f"{len(data)} unique parts").alignment = left
    if len(data) > 0:
        ws.cell(row=tr, column=4,
                value=f"=SUM(D4:D{tr - 1})").alignment = center
    ws.row_dimensions[tr].height = 22

    # ── Freeze header rows ───────────────────────────────────
    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"summary_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export_summary_pdf/<shipment_ids>')
def export_summary_pdf(shipment_ids):
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')

    ids = [i.strip() for i in shipment_ids.split(',') if i.strip()]

    conn = get_db()
    cur = dict_cursor(conn)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            leftMargin=36, rightMargin=36,
                            topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = []

    title_style = styles['Title']
    elements.append(Paragraph("QAC Discussion Summary Report", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    elements.append(Spacer(1, 14))

    for sid in ids:
        cur.execute("""
            SELECT "Part_Name", "Part_Number", "Model", status, "PIC", "Supplier_name", "Date_sent", dealer_id
            FROM shipments WHERE "Sl_No"=%s
        """, (sid,))
        ship = cur.fetchone()
        if not ship:
            continue

        part_name      = ship['Part_Name'] or ''
        part_number    = ship['Part_Number'] or ''
        model          = ship['Model'] or ''
        final_status   = ship['status'] or 'Open'
        pi_number_ship = ship['PIC'] or ''
        supplier       = ship['Supplier_name'] or ''
        date_sent      = ship['Date_sent'] or ''
        dealer_id_val  = ship['dealer_id'] or ''

        elements.append(Paragraph(
            f"<b>Part Name:</b> {part_name} &nbsp;&nbsp; <b>Part Number:</b> {part_number}",
            styles['Heading2']
        ))
        elements.append(Paragraph(
            f"<b>Model:</b> {model} &nbsp;&nbsp; <b>Supplier:</b> {supplier} &nbsp;&nbsp; "
            f"<b>Date:</b> {date_sent} &nbsp;&nbsp; <b>PI Number:</b> {pi_number_ship} &nbsp;&nbsp; "
            f"<b>Dealer ID:</b> {dealer_id_val}",
            styles['Normal']
        ))
        elements.append(Spacer(1, 8))

        cur.execute("""
            SELECT pi_number, message, created_at,
                   author_name, author_username, author_role,
                   edited, edited_at, edited_by
            FROM "Discussion"
            WHERE shipment_id=%s AND is_deleted=0
            ORDER BY "Sl_No" ASC
        """, (sid,))
        discussions = cur.fetchall()

        role_dept_map = {
            'QACAdmin': 'QAC', 'QICAdmin': 'QIC', 'ProductionAdmin': 'Production',
            'QAC': 'QAC', 'QIC': 'QIC', 'Production': 'Production',
            'Dealer': 'Dealer', 'Supplier': 'Supplier', 'General': 'TSD',
        }

        if discussions:
            table_data = [[
                Paragraph('<b>PI Number</b>', styles['Normal']),
                Paragraph('<b>Message</b>', styles['Normal']),
                Paragraph('<b>Date &amp; Time</b>', styles['Normal']),
                Paragraph('<b>Admin/User</b>', styles['Normal']),
                Paragraph('<b>Discussed By</b>', styles['Normal']),
            ]]

            for d in discussions:
                role = d['author_role'] or ''
                admin_or_user = 'Admin' if role in ('QACAdmin', 'QICAdmin', 'ProductionAdmin') else 'User'
                dept_label = role_dept_map.get(role, role) if role else ''
                uname = d['author_username'] or ''
                discussed_by = f"{uname} ({dept_label})" if dept_label else uname
                table_data.append([
                    Paragraph(str(d['pi_number'] or ''), styles['Normal']),
                    Paragraph(str(d['message'] or ''), styles['Normal']),
                    Paragraph(str(d['created_at'] or ''), styles['Normal']),
                    Paragraph(admin_or_user, styles['Normal']),
                    Paragraph(discussed_by, styles['Normal']),
                ])

            col_widths_pdf = [55, 210, 90, 50, 165]
            tbl = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#32a7b4')),
                ('TEXTCOLOR',  (0, 0), (-1, 0), colors.HexColor('#d4feff')),
                ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0, 0), (-1, 0), 7),
                ('FONTSIZE',   (0, 1), (-1, -1), 7),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#a4bcff'), colors.white]),
                ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#babad6')),
                ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING',  (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING',   (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
            ]))
            elements.append(tbl)
        else:
            elements.append(Paragraph("No messages in this discussion.", styles['Normal']))

        elements.append(Spacer(1, 6))

        status_color = '#333333'
        fs_lower = final_status.lower()
        if 'closed' in fs_lower:
            status_color = '#c0392b'
        elif 'tsd' in fs_lower:
            status_color = '#b06820'
        elif 'inprogress' in fs_lower or 'in progress' in fs_lower:
            status_color = '#b06820'
        elif 'open' in fs_lower:
            status_color = '#1a8a5a'

        elements.append(Paragraph(
            f'<b>Final Result / Status:</b> <font color="{status_color}"><b>{final_status}</b></font>',
            styles['Normal']
        ))
        elements.append(Spacer(1, 18))
        elements.append(Paragraph('<hr/>', styles['Normal']))
        elements.append(Spacer(1, 10))

    conn.close()
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer,
                     as_attachment=True,
                     download_name="summary_report.pdf",
                     mimetype='application/pdf')


@app.route('/summary_dashboard')
def summary_dashboard():
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')

    vals = {
        'part_name':   request.args.get('part_name', ''),
        'model':       request.args.get('model', ''),
        'total_cases': request.args.get('total_cases', ''),
        'status':      request.args.get('status', '')
    }

    data = get_summary_data(vals)

    # Build filter query string for the Excel export link (pass current filters)
    export_qs = '&'.join(f"{k}={v}" for k, v in vals.items() if v)
    export_url = f"/export_summary_excel?{export_qs}" if export_qs else "/export_summary_excel"

    rows = ""
    for d in data:
        status_cls = get_status_css_class(d['latest_status'] or 'Open')
        rows += f"""
        <tr>
          <td><strong style="color:var(--steel-frost)">{d['Part_Name']}</strong></td>
          <td>{d['Model']}</td>
          <td style="font-weight:bold; color:var(--accent-gold); text-align:center;">{d['total_cases']}</td>
          <td><span class="{status_cls}">{d['latest_status'] or 'Open'}</span></td>
          <td>
            <a href="/export_summary_pdf/{d['shipment_ids']}" class="btn btn-sm btn-success">
              &#128196; Export Report
            </a>
          </td>
        </tr>"""

    status_options = ['Open', 'Inprogress', 'Inprogress : TSD', 'Closed']
    status_opts = '<option value="">All</option>'
    for v in status_options:
        sel_attr = 'selected' if vals.get('status') == v else ''
        status_opts += f'<option value="{v}" {sel_attr}>{v}</option>'

    is_admin = 'admin' in session
    header_fn = render_admin_header if is_admin else render_dealer_header

    html = get_base_style() + header_fn() + f"""
    <div class="page-wrap">

      <!-- Top bar: title + Export All button -->
      <div class="card" style="display:flex; justify-content:space-between; align-items:center;
           padding:14px 22px; margin-bottom:0; border-bottom: 1px solid var(--border-light);
           border-radius:8px 8px 0 0;">
        <h3 style="margin:0; font-family:'Rajdhani',sans-serif; color:var(--steel-ice);">
          &#128202; Summary Dashboard
          <small style="color:var(--text-dim); font-weight:400; font-family:'IBM Plex Sans',sans-serif; font-size:.82rem;">
            ({len(data)} groups)
          </small>
        </h3>
        <a href="{export_url}" class="btn btn-success" style="text-decoration:none; display:inline-flex; align-items:center; gap:6px;">
          &#11015;&nbsp;Export All to Excel
        </a>
      </div>

      <!-- Filter form -->
      <div class="card" style="border-radius:0 0 0 0; border-top:none; margin-bottom:0; padding-top:14px;">
        <form method="get" action="/summary_dashboard" class="form-row">
          <div class="form-group">
            <label>Part Name</label>
            <input type="text" name="part_name" value="{vals['part_name']}" placeholder="Part Name" style="min-width:160px">
          </div>
          <div class="form-group">
            <label>Model</label>
            <input type="text" name="model" value="{vals['model']}" placeholder="Model">
          </div>
          <div class="form-group">
            <label>Total Cases</label>
            <input type="text" name="total_cases" value="{vals['total_cases']}" placeholder="Count">
          </div>
          <div class="form-group">
            <label>Status</label>
            <select name="status">{status_opts}</select>
          </div>
          <div class="form-group">
            <label>&nbsp;</label>
            <div style="display: flex; gap: 8px;">
              <button type="submit" class="btn btn-primary" style="height:36px;">&#128269; Filter</button>
              <a href="/summary_dashboard" class="btn" style="background:rgba(0,0,0,0.1); color:var(--text-dim);
                 text-decoration:none; height:36px; display:inline-flex; align-items:center;
                 border:1px solid var(--border-light); border-radius:4px; font-size:.85rem; padding:0 12px;">Reset</a>
            </div>
          </div>
        </form>
      </div>

      <!-- Table -->
      <div class="card" style="border-radius:0 0 8px 8px; border-top:none; margin-bottom:18px;">
        <div class="table-wrap">
          <table>
            <thead>
              <tr>
                <th>Part Name</th>
                <th>Model</th>
                <th style="text-align:center;">Total Cases (Qty)</th>
                <th>Latest Status</th>
                <th>Individual Report</th>
              </tr>
            </thead>
            <tbody>{rows if rows else '<tr><td colspan="5" style="text-align:center; padding:20px; color:var(--text-muted)">No matching records found.</td></tr>'}</tbody>
          </table>
        </div>
      </div>
    </div>

    <div class="view-switcher">
      <a href="/admin_dashboard" class="vs-normal"><span class="vs-icon">&#9776;</span> Normal View</a>
      <a href="/summary_dashboard" class="vs-summary vs-active"><span class="vs-icon">&#128202;</span> Summary View</a>
    </div>
    """
    return render_template_string(html)


# =========================================================
# DELETE / TRASH
# =========================================================
@app.route('/delete_shipment/<int:slno>')
def delete_shipment(slno):
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')
    conn = get_db()
    cur = conn.cursor()
    cur.execute('UPDATE shipments SET is_deleted=1 WHERE "Sl_No"=%s', (slno,))
    conn.commit()
    conn.close()
    return redirect('/admin_dashboard' if 'admin' in session else '/dealer_dashboard')


@app.route('/trash')
def trash():
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')
    is_admin = 'admin' in session
    page = int(request.args.get('page', 1))
    offset = (page - 1) * PER_PAGE
    conn = get_db()
    cur = dict_cursor(conn)
    cur.execute('SELECT COUNT(*) as cnt FROM shipments WHERE is_deleted=1')
    total_items = cur.fetchone()['cnt']
    total_pages = max(1, (total_items + PER_PAGE - 1) // PER_PAGE)
    cur.execute('SELECT * FROM shipments WHERE is_deleted=1 ORDER BY "Sl_No" DESC LIMIT %s OFFSET %s', (PER_PAGE, offset))
    shipments = cur.fetchall()
    conn.close()
    rows = ""
    display_sl_start = total_items - offset
    for idx, s in enumerate(shipments):
        display_sl = display_sl_start - idx
        status_cls = get_status_css_class(s['status'] or 'Open')
        if is_admin:
            action_btns = f"""
            <a href="{url_for('restore_shipment', slno=s['Sl_No'])}" class="btn btn-sm btn-success">&#8635; Restore</a>
            <a href="{url_for('permanent_delete_shipment', slno=s['Sl_No'])}" class="btn btn-sm btn-danger"
               onclick="return confirm('Permanently delete? This cannot be undone.')">&#128465; Delete Forever</a>"""
        else:
            action_btns = f'<a href="{url_for("restore_shipment", slno=s["Sl_No"])}" class="btn btn-sm btn-success">&#8635; Restore</a>'
        rows += f"""<tr>
          <td style="color:var(--text-muted)">{display_sl}</td>
          <td>{s['Date_sent'] or ''}</td><td>{s['Model'] or ''}</td>
          <td><code>{s['Part_Number'] or ''}</code></td>
          <td><strong style="color:var(--steel-frost)">{s['Part_Name'] or ''}</strong></td>
          <td>{s['Supplier_name'] or ''}</td>
          <td><span class="{status_cls}">{s['status'] or ''}</span></td>
          <td style="color:var(--accent-gold);font-weight:600">{s['dealer_id'] or ''}</td>
          <td>{action_btns}</td></tr>"""
    pagination = "".join([f'<a href="?page={p}" class="{"active" if p==page else ""}">{p}</a>' for p in range(1, total_pages+1)])
    back_link = '/admin_dashboard' if is_admin else '/dealer_dashboard'
    header_html = render_admin_header() if is_admin else render_dealer_header()
    html = get_base_style() + header_html + f"""
    <div class="page-wrap">
      <div style="margin-bottom:12px">
        <a href="{back_link}" class="btn btn-primary">&#8592; Back to Dashboard</a>
      </div>
      <div class="card">
        <h2>&#128465;&#65039; Trash &nbsp;<small style="color:var(--text-dim);font-weight:400;font-family:'IBM Plex Sans',sans-serif;font-size:.85rem">({total_items} items)</small></h2>
        <p style="color:var(--text-muted);font-size:0.82rem;margin-top:6px;margin-bottom:14px">
          Items here have been soft-deleted.
          {"Admins can restore or permanently delete. Users can restore only." if is_admin else "You can restore items back to the dashboard."}
        </p>
        <div class="table-wrap"><table>
          <thead><tr><th>#</th><th>Date</th><th>Model</th><th>Part No.</th><th>Part Name</th><th>Supplier</th><th>Status</th><th>Dealer ID</th><th>Actions</th></tr></thead>
          <tbody>{rows}</tbody>
        </table></div>
        <div class="pagination">{pagination}</div>
      </div>
    </div>"""
    return render_template_string(html)


@app.route('/restore_shipment/<int:slno>')
def restore_shipment(slno):
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')
    conn = get_db()
    conn.cursor().execute('UPDATE shipments SET is_deleted=0 WHERE "Sl_No"=%s', (slno,))
    conn.commit()
    conn.close()
    return redirect('/trash')


@app.route('/permanent_delete_shipment/<int:slno>')
def permanent_delete_shipment(slno):
    if 'admin' not in session:
        return redirect('/admin_login')
    conn = get_db()
    conn.cursor().execute('DELETE FROM shipments WHERE "Sl_No"=%s', (slno,))
    conn.commit()
    conn.close()
    return redirect('/trash')


# =========================================================
# UPDATE ONEDRIVE URL
# =========================================================
@app.route('/update_onedrive/<int:shipment_id>', methods=['POST'])
def update_onedrive(shipment_id):
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')
    onedrive_url = request.form.get('onedrive_url', '').strip()
    conn = get_db()
    conn.cursor().execute(
        'UPDATE shipments SET onedrive_url=%s WHERE "Sl_No"=%s',
        (onedrive_url, shipment_id)
    )
    conn.commit()
    conn.close()
    return redirect(url_for('view_discussion', shipment_id=shipment_id))


# =========================================================
# DISCUSSION
# =========================================================
def _tsd_auto_status(current_status):
    if current_status in ('Open', 'Inprogress'):
        return 'Inprogress : TSD'
    return current_status


@app.route('/discussion/<int:shipment_id>', methods=['GET', 'POST'])
def view_discussion(shipment_id):
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')
    name, username, role = get_current_user()
    conn = get_db()
    cur = dict_cursor(conn)

    if request.method == 'POST':
        action = request.form.get('action', 'post')

        if action == 'post':
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cur.execute("""INSERT INTO "Discussion"
            (shipment_id,pi_number,message,dept,created_at,author_name,author_username,author_role)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
            (shipment_id, request.form.get('pi_number',''), request.form.get('message',''),
             role, now, name, username, role))

            is_tsd = ('dealer' in session and
                      session.get('dealer_department', '') == 'General')
            if is_tsd:
                cur.execute('SELECT status FROM shipments WHERE "Sl_No"=%s', (shipment_id,))
                row = cur.fetchone()
                if row:
                    new_status = _tsd_auto_status(row['status'] or 'Open')
                    cur.execute(
                        'UPDATE shipments SET status=%s WHERE "Sl_No"=%s',
                        (new_status, shipment_id)
                    )

        elif action == 'edit':
            disc_id = int(request.form.get('disc_id', 0))
            cur.execute('SELECT * FROM "Discussion" WHERE "Sl_No"=%s', (disc_id,))
            disc = cur.fetchone()
            if disc and (disc['author_username'] == username or 'admin' in session):
                cur.execute('UPDATE "Discussion" SET message=%s,edited=1,edited_at=%s,edited_by=%s WHERE "Sl_No"=%s',
                    (request.form.get('message',''), datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     f"{name} ({username})", disc_id))

        elif action == 'delete':
            disc_id = int(request.form.get('disc_id', 0))
            cur.execute('SELECT * FROM "Discussion" WHERE "Sl_No"=%s', (disc_id,))
            disc = cur.fetchone()
            if disc and (disc['author_username'] == username or 'admin' in session):
                cur.execute('UPDATE "Discussion" SET is_deleted=1,edited_by=%s,edited_at=%s WHERE "Sl_No"=%s',
                    (f"Deleted by {name} ({username})", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), disc_id))

        conn.commit()
        conn.close()
        return redirect(url_for('view_discussion', shipment_id=shipment_id))

    cur.execute('SELECT * FROM shipments WHERE "Sl_No"=%s', (shipment_id,))
    shipment = cur.fetchone()
    cur.execute('SELECT * FROM "Discussion" WHERE shipment_id=%s ORDER BY "Sl_No" ASC', (shipment_id,))
    discussions = cur.fetchall()
    conn.close()
    back_link = '/admin_dashboard' if 'admin' in session else '/dealer_dashboard'

    current_onedrive = ''
    if shipment:
        try:
            current_onedrive = shipment['onedrive_url'] or ''
        except Exception:
            current_onedrive = ''

    disc_html = ""
    for d in discussions:
        is_mine = (d['author_username'] == username)
        can_edit = is_mine or 'admin' in session
        deleted = d['is_deleted'] == 1
        role_cls = ''.join(c for c in (d['author_role'] or 'user') if c.isalnum())
        author_role = d['author_role'] or ''
        is_admin_role = author_role in ('QACAdmin', 'QICAdmin', 'ProductionAdmin')
        side = 'left' if is_admin_role else 'right'
        edited_tag = f'<span class="edited-tag">&#9998; edited {d["edited_at"] or ""}</span>' if d['edited'] and not deleted else ""

        display_role = 'TSD' if author_role == 'General' else author_role

        if deleted:
            bubble_extra = 'deleted-bubble'
            msg_content = "&#128465; This message was deleted"
            deleted_by = f'<div class="bubble-time"><span style="color:#f07070;font-size:.69rem">{d["edited_by"] or "Deleted"} &bull; {d["edited_at"] or ""}</span></div>'
            actions_html = ""
            edit_form = ""
        else:
            bubble_extra = ''
            msg_content = d['message'] or ''
            deleted_by = ""
            edit_form = ""
            if can_edit:
                edit_form = f"""
                <div id="ef-{d['Sl_No']}" class="bubble-edit-form" style="display:none">
                  <form method="POST">
                    <input type="hidden" name="action" value="edit">
                    <input type="hidden" name="disc_id" value="{d['Sl_No']}">
                    <textarea name="message" rows="2" style="width:100%">{d['message']}</textarea><br>
                    <button type="submit" class="btn btn-sm btn-primary" style="margin-top:4px">Save</button>
                    <button type="button" class="btn btn-sm" style="background:rgba(42,106,191,0.15);color:var(--chrome-2);border:1px solid var(--border-light);margin-top:4px"
                      onclick="document.getElementById('ef-{d['Sl_No']}').style.display='none'">Cancel</button>
                  </form>
                </div>"""
                actions_html = f"""
                <div class="bubble-actions">
                  <button type="button" class="btn btn-sm btn-warn"
                    onclick="document.getElementById('ef-{d["Sl_No"]}').style.display='block'">&#9998; Edit</button>
                  <form method="POST" style="display:inline">
                    <input type="hidden" name="action" value="delete">
                    <input type="hidden" name="disc_id" value="{d['Sl_No']}">
                    <button type="submit" class="btn btn-sm btn-danger"
                      onclick="return confirm('Delete this message?')">&#128465; Delete</button>
                  </form>
                </div>"""
            else:
                actions_html = ""

        pi_badge = f'<span class="bubble-pi">PI: {d["pi_number"]}</span>' if d['pi_number'] else ''
        disc_html += f"""
        <div class="chat-bubble-wrap {side}">
          <div class="chat-bubble {side} {bubble_extra}">
            <div class="bubble-header">
              <span class="bubble-author">{d['author_name'] or 'Unknown'}</span>
              <span class="role-tag role-{role_cls}">{display_role}</span>
              {pi_badge}
            </div>
            <div style="line-height:1.55">{msg_content}</div>
            {deleted_by}
            <div class="bubble-time">
              <span>{'@' + (d['author_username'] or '')}</span>
              <span>&bull;</span>
              <span>{d['created_at'] or ''}</span>
              {edited_tag}
            </div>
            {actions_html}
            {edit_form}
          </div>
        </div>"""

    if not disc_html:
        disc_html = "<p style='color:var(--text-dim);text-align:center;padding:32px'>No messages yet. Be the first to post!</p>"

    part_info = ""
    if shipment:
        current_status = shipment['status'] or 'Open'
        status_cls = get_status_css_class(current_status)
        dealer_id_display = shipment.get('dealer_id') or ''
        part_info = f"""<div class="alert alert-info" style="margin-bottom:14px">
          <strong>Part:</strong> {shipment['Part_Name']} &nbsp;|&nbsp;
          <strong>No.:</strong> {shipment['Part_Number']} &nbsp;|&nbsp;
          <strong>Model:</strong> {shipment['Model']} &nbsp;|&nbsp;
          <strong>Dealer ID:</strong> <span style="color:var(--accent-gold);font-weight:700">{dealer_id_display}</span> &nbsp;|&nbsp;
          <strong>Status:</strong> <span class="{status_cls}">{current_status}</span>
        </div>"""

    onedrive_link_btn = ""
    if current_onedrive:
        onedrive_link_btn = f'<a href="{current_onedrive}" target="_blank" class="btn btn-success btn-sm">&#128279; Open OneDrive</a>'

    onedrive_bar = f"""
    <div class="onedrive-bar">
      <label>&#9729;&#65039; OneDrive URL</label>
      <form method="POST" action="/update_onedrive/{shipment_id}" style="display:flex;gap:8px;flex:1;align-items:center;flex-wrap:wrap">
        <input type="text" name="onedrive_url" value="{current_onedrive}"
               placeholder="Paste OneDrive / SharePoint URL here..." style="flex:1;min-width:260px">
        <button type="submit" class="btn btn-primary btn-sm">&#128190; Save URL</button>
      </form>
      {onedrive_link_btn}
    </div>"""

    msg_count = len(discussions)
    is_admin = 'admin' in session
    header_fn = render_admin_header if is_admin else render_dealer_header

    html = get_base_style() + header_fn() + f"""
    <div class="page-wrap">
      <div style="margin-bottom:12px">
        <a href="/export_discussion/{shipment_id}" class="btn btn-success">&#11015; Export to Excel</a>
        <a href="{back_link}" class="btn btn-primary">&#8592; Back to Dashboard</a>
      </div>
      {part_info}
      {onedrive_bar}
      <div class="card">
        <h2>&#128172; Post a Message</h2>
        <form method="POST" style="margin-top:12px">
          <input type="hidden" name="action" value="post">
          <div class="form-row">
            <div class="form-group"><label>PI Number</label>
              <input type="text" name="pi_number" placeholder="PI Number"></div>
          </div>
          <div class="form-group" style="margin-top:10px"><label>Message</label>
            <textarea name="message" rows="3" style="width:100%" required placeholder="Type your message..."></textarea></div>
          <button type="submit" class="btn btn-primary" style="margin-top:10px">&#128228; Post Message</button>
        </form>
      </div>
      <div class="card">
        <h3>Messages ({msg_count})</h3>
        <div class="chat-container" style="margin-top:10px" id="chatbox">{disc_html}</div>
      </div>
    </div>
    <div class="view-switcher">
      <a href="/admin_dashboard" class="vs-normal"><span class="vs-icon">&#9776;</span> Normal View</a>
      <a href="/summary_dashboard" class="vs-summary vs-active"><span class="vs-icon">&#128202;</span> Summary View</a>
    </div>
    <script>
      var chatbox = document.getElementById('chatbox');
      if(chatbox) chatbox.scrollTop = chatbox.scrollHeight;
    </script>"""
    return render_template_string(html)


# =========================================================
# MANAGE ADMINS
# =========================================================
@app.route('/manage_admins', methods=['GET', 'POST'])
def manage_admins():
    if 'admin' not in session:
        return redirect('/admin_login')
    conn = get_db()
    cur = dict_cursor(conn)
    msg = ""
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            uname = request.form.get('username','').strip()
            cur.execute("SELECT COUNT(*) as cnt FROM admins WHERE username=%s", (uname,))
            if cur.fetchone()['cnt'] > 0:
                msg = "add_err"
            else:
                try:
                    cur.execute("INSERT INTO admins(name,username,password,role) VALUES(%s,%s,%s,%s)",
                        (request.form['name'], uname, request.form['password'], request.form['role']))
                    conn.commit()
                    msg = "add_ok"
                except Exception:
                    conn.rollback()
                    msg = "add_err"
        elif action == 'delete':
            aid = int(request.form.get('admin_id', 0))
            if aid != session['admin']:
                cur.execute('DELETE FROM admins WHERE "Sl_No"=%s', (aid,))
                conn.commit()
                msg = "del_ok"
            else:
                msg = "del_self"
        elif action == 'change_password':
            cur.execute('UPDATE admins SET password=%s WHERE "Sl_No"=%s',
                (request.form['new_password'], int(request.form.get('admin_id', 0))))
            conn.commit()
            msg = "pw_ok"
    cur.execute('SELECT * FROM admins ORDER BY "Sl_No"')
    admins = cur.fetchall()
    conn.close()
    alert_map = {
        'add_ok':   ('success', 'Admin added successfully.'),
        'add_err':  ('danger',  'Username already exists.'),
        'del_ok':   ('success', 'Admin deleted.'),
        'del_self': ('danger',  'Cannot delete your own account.'),
        'pw_ok':    ('success', 'Password updated.'),
    }
    alert_html = ""
    if msg in alert_map:
        cls, text = alert_map[msg]
        alert_html = f'<div class="alert alert-{cls}">{text}</div>'
    rows = ""
    for a in admins:
        role_cls = ''.join(c for c in a['role'] if c.isalnum())
        you = " <small style='color:var(--accent-gold)'>(you)</small>" if a['Sl_No'] == session['admin'] else ""
        rows += f"""<tr>
          <td style="color:var(--text-muted)">{a['Sl_No']}</td>
          <td>{a['name']}{you}</td>
          <td><code>{a['username']}</code></td>
          <td><span class="role-tag role-{role_cls}">{a['role']}</span></td>
          <td>
            <form method="POST" style="display:inline">
              <input type="hidden" name="action" value="change_password">
              <input type="hidden" name="admin_id" value="{a['Sl_No']}">
              <input type="password" name="new_password" placeholder="New password" style="width:120px">
              <button type="submit" class="btn btn-sm btn-warn">Update PW</button>
            </form>
            {"" if a['Sl_No'] == session['admin'] else f"""
            <form method="POST" style="display:inline">
              <input type="hidden" name="action" value="delete">
              <input type="hidden" name="admin_id" value="{a['Sl_No']}">
              <button type="submit" class="btn btn-sm btn-danger" onclick="return confirm('Delete admin?')">Delete</button>
            </form>"""}
          </td></tr>"""
    role_opts = "".join([f'<option value="{r}">{r}</option>' for r in ['QACAdmin','QICAdmin','ProductionAdmin']])
    html = get_base_style() + render_admin_header() + f"""
    <div class="page-wrap">
      {alert_html}
      <div class="card">
        <h2>&#10133; Add New Admin</h2><br>
        <form method="POST" class="form-row">
          <input type="hidden" name="action" value="add">
          <div class="form-group"><label>Full Name</label>
            <input type="text" name="name" required placeholder="Admin full name"></div>
          <div class="form-group"><label>Username</label>
            <input type="text" name="username" required placeholder="Username"></div>
          <div class="form-group"><label>Password</label>
            <input type="password" name="password" required placeholder="Password"></div>
          <div class="form-group"><label>Role</label>
            <select name="role">{role_opts}</select></div>
          <div class="form-group"><label>&nbsp;</label>
            <button type="submit" class="btn btn-primary">Add Admin</button></div>
        </form>
      </div>
      <div class="card">
        <h2>&#128101; All Admins</h2>
        <div class="table-wrap"><table>
          <thead><tr><th>ID</th><th>Name</th><th>Username</th><th>Role</th><th>Actions</th></tr></thead>
          <tbody>{rows}</tbody>
        </table></div>
      </div>
    </div>"""
    return render_template_string(html)


# =========================================================
# MANAGE USERS
# =========================================================
@app.route('/manage_users', methods=['GET', 'POST'])
def manage_users():
    if 'admin' not in session:
        return redirect('/admin_login')
    conn = get_db()
    cur = dict_cursor(conn)
    msg = ""
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'delete':
            cur.execute('DELETE FROM dealers WHERE "Sl_No"=%s', (int(request.form.get('user_id', 0)),))
            conn.commit()
            msg = "del_ok"
        elif action == 'change_password':
            cur.execute('UPDATE dealers SET password=%s WHERE "Sl_No"=%s',
                (request.form['new_password'], int(request.form.get('user_id', 0))))
            conn.commit()
            msg = "pw_ok"
    cur.execute('SELECT * FROM dealers ORDER BY "Sl_No"')
    users = cur.fetchall()
    conn.close()
    alert_html = ""
    if msg == 'del_ok':
        alert_html = '<div class="alert alert-success">User deleted.</div>'
    elif msg == 'pw_ok':
        alert_html = '<div class="alert alert-success">Password updated.</div>'
    rows = ""
    for u in users:
        dept_display = 'TSD' if u['department'] == 'General' else u['department']
        rows += f"""<tr>
          <td style="color:var(--text-muted)">{u['Sl_No']}</td>
          <td>{u['name']}</td><td><code>{u['username']}</code></td><td>{dept_display}</td>
          <td>
            <form method="POST" style="display:inline">
              <input type="hidden" name="action" value="change_password">
              <input type="hidden" name="user_id" value="{u['Sl_No']}">
              <input type="password" name="new_password" placeholder="New password" style="width:120px">
              <button type="submit" class="btn btn-sm btn-warn">Update PW</button>
            </form>
            <form method="POST" style="display:inline">
              <input type="hidden" name="action" value="delete">
              <input type="hidden" name="user_id" value="{u['Sl_No']}">
              <button type="submit" class="btn btn-sm btn-danger" onclick="return confirm('Delete user?')">Delete</button>
            </form>
          </td></tr>"""
    html = get_base_style() + render_admin_header() + f"""
    <div class="page-wrap">
      {alert_html}
      <div class="card">
        <h2>&#128100; All Users ({len(users)})</h2>
        <div class="table-wrap"><table>
          <thead><tr><th>ID</th><th>Name</th><th>Username</th><th>Department</th><th>Actions</th></tr></thead>
          <tbody>{rows}</tbody>
        </table></div>
      </div>
    </div>"""
    return render_template_string(html)


# =========================================================
# REPORTS  — accessible to BOTH admins and users/dealers
# =========================================================
@app.route('/reports')
def reports():
    if 'admin' not in session and 'dealer' not in session:
        return redirect('/')

    conn = get_db()
    cur = dict_cursor(conn)

    # Category data — skip nulls/empty
    cur.execute("""
        SELECT category, COUNT(*) as count
        FROM shipments
        WHERE is_deleted = 0
          AND category IS NOT NULL
          AND category != ''
        GROUP BY category
    """)
    cat_res = cur.fetchall()

    # Status data
    cur.execute("""
        SELECT status, COUNT(*) as count
        FROM shipments
        WHERE is_deleted = 0
        GROUP BY status
    """)
    stat_res = cur.fetchall()

    # Weekly trend
    cur.execute("""
        SELECT TO_CHAR("Date_sent"::DATE, 'WW') as week, COUNT(*) as count
        FROM shipments
        WHERE is_deleted = 0
        GROUP BY week
        ORDER BY week ASC
    """)
    trend_res = cur.fetchall()
    conn.close()

    is_admin = 'admin' in session
    header_fn = render_admin_header if is_admin else render_dealer_header

    return header_fn() + get_base_style() + render_template_string(
        REPORT_HTML,
        cat_data=json.dumps(cat_res),
        stat_data=json.dumps(stat_res),
        trend_data=json.dumps(trend_res)
    )


# =========================================================
# MAIN
# =========================================================
if __name__ == "__main__":
    import socket
    init_db()
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()
    except:
        local_ip = "127.0.0.1"
    port = 5000
    print("=" * 60)
    print("  QAC Discussion Portal by Toyota Kirloskar Motors")
    print("=" * 60)
    print(f"  Local:         http://127.0.0.1:{port}")
    print(f"  Network (LAN): http://{local_ip}:{port}")
    print()
    print("  Default Admin Credentials:")
    print("  Role             | Username  | Password")
    print("  QACAdmin         | qacadmin  | qac123")
    print("  QICAdmin         | qicadmin  | qic123")
    print("  ProductionAdmin  | prodadmin | prod123")
    print("=" * 60)
    app.run(host='0.0.0.0', port=port, debug=False)
